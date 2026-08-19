"""Loads the real product catalog (Product.xlsm, sheet "Export") and
imports it into the existing `Item` table.

This is a mechanical preprocessing step - forward-filling ItemFamily is
pure pandas/openpyxl bookkeeping, never fuzzy matching or an LLM. The
original ItemDescription is never modified; only a derived
normalized_description is added alongside it for matching.
"""
from __future__ import annotations

from dataclasses import dataclass

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Item
from app.services.scripted.normalization import normalize_item_text

SHEET_NAME = "Export"
REQUIRED_COLUMNS = ("ItemFamily", "ItemNumber", "ItemDescription")


class CatalogLoadError(Exception):
    pass


@dataclass
class CatalogRow:
    item_number: str
    original_description: str
    normalized_description: str
    item_family: str | None


def _header_index(header_row) -> dict[str, int]:
    idx = {cell: i for i, cell in enumerate(header_row) if cell is not None}
    missing = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing:
        raise CatalogLoadError(
            f"Product.xlsm/{SHEET_NAME} is missing required column(s): "
            f"{', '.join(missing)}")
    return idx


def load_catalog(path: str) -> list[CatalogRow]:
    """Load + forward-fill + normalize the catalog. Rows missing an
    ItemNumber or ItemDescription (blank trailing rows, filter-summary
    rows like the real file's "No filters applied" footer) are skipped -
    they carry no catalog information, not a data error.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise CatalogLoadError(
            f"Product.xlsm has no sheet named {SHEET_NAME!r} "
            f"(found: {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise CatalogLoadError(f"{SHEET_NAME} sheet is empty") from None
    idx = _header_index(header)
    fam_i, nb_i, desc_i = (idx["ItemFamily"], idx["ItemNumber"],
                           idx["ItemDescription"])

    catalog: list[CatalogRow] = []
    current_family: str | None = None
    for row in rows_iter:
        family_cell = row[fam_i] if fam_i < len(row) else None
        if family_cell:
            current_family = str(family_cell).strip()

        item_number = row[nb_i] if nb_i < len(row) else None
        description = row[desc_i] if desc_i < len(row) else None
        if not item_number or not description:
            continue  # blank/footer row - no catalog information

        item_number = str(item_number).strip()
        description = str(description).strip()
        catalog.append(CatalogRow(
            item_number=item_number,
            original_description=description,
            normalized_description=normalize_item_text(description),
            item_family=current_family))
    return catalog


def duplicate_descriptions(catalog: list[CatalogRow]) -> dict[str, list[str]]:
    """original_description -> [item_number, ...] for every description
    shared by more than one item_number - the set match_item.py must never
    silently auto-resolve (spec section 26)."""
    by_desc: dict[str, list[str]] = {}
    for row in catalog:
        by_desc.setdefault(row.original_description, []).append(row.item_number)
    return {desc: nbs for desc, nbs in by_desc.items() if len(nbs) > 1}


def import_catalog(session: Session, path: str) -> int:
    """Idempotent upsert of the catalog into the existing `Item` table
    (item_number / item_desc=original description / category=item_family).
    Existing item_number rows are updated in place rather than duplicated;
    unit_price (not present in Product.xlsm) is left untouched on update
    and null on insert. Returns the number of rows processed.
    """
    catalog = load_catalog(path)
    existing = {i.item_number: i for i in session.scalars(select(Item))}
    for row in catalog:
        item = existing.get(row.item_number)
        if item is None:
            item = Item(item_number=row.item_number,
                        item_desc=row.original_description,
                        category=row.item_family or "Uncategorized")
            session.add(item)
            existing[row.item_number] = item
        else:
            item.item_desc = row.original_description
            item.category = row.item_family or item.category
    session.flush()
    return len(catalog)
